"""
vowelfreq.py — standalone vowel-frequency pattern analyzer for any wordlist.

Run against a plaintext wordlist (one word per line) and it produces an
Excel workbook matching the format used on vowelfreq.com:

    Sheet "Vowel Analysis":
        Row 1: Word Length | 3 characters | 4 characters | ...
        Row 2: Total Words | <count>      | <count>      | ...
        Row 3+: <# of vowels> | <% for length 3> | <% for length 4> | ...

    Sheet "Length_3", "Length_4", ..., one per length in range:
        Pattern | Count | Percentage
        (patterns use ?1 for vowels, ?2 for non-vowels — Hashcat custom charset
         notation, sorted by Count descending)

A sibling CSV (<stem>_vowel_counts.csv) is also written for offline use.

Usage:
    python3 vowelfreq.py --input words.txt
    python3 vowelfreq.py --input rockyou.txt --min 4 --max 16 --output out.xlsx
"""

import argparse
import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font

VOWELS = set("aeiou")


def parse_args():
    p = argparse.ArgumentParser(
        description="Vowel-frequency pattern analyzer for wordlists "
                    "(produces an xlsx compatible with vowelfreq.com).",
    )
    p.add_argument("--input", required=True,
                   help="Path to a wordlist (one word per line).")
    p.add_argument("--min", type=int, default=3,
                   help="Minimum word length to include (default: 3).")
    p.add_argument("--max", type=int, default=20,
                   help="Maximum word length to include (default: 20).")
    p.add_argument("--output", default="pattern_results.xlsx",
                   help="Output xlsx path (default: pattern_results.xlsx).")
    args = p.parse_args()
    if args.min > args.max:
        p.error("--min cannot be greater than --max.")
    if args.min < 1:
        p.error("--min must be at least 1.")
    return args


def load_words(path, min_len, max_len):
    """Return words filtered to [min_len, max_len], lowercased, alpha-only."""
    kept = []
    skipped = 0
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            word = line.strip().lower()
            if not word:
                continue
            if not (min_len <= len(word) <= max_len):
                skipped += 1
                continue
            if not word.isalpha():
                skipped += 1
                continue
            kept.append(word)
    return kept, skipped


def analyze(words):
    """Group counts by length and by pattern — single pass over input."""
    by_length = {}           # length -> Counter of vowel_count -> word_count
    patterns_by_length = {}  # length -> Counter of ?1/?2 pattern -> word_count
    for word in words:
        length = len(word)
        n_vowels = sum(1 for c in word if c in VOWELS)
        by_length.setdefault(length, Counter())[n_vowels] += 1
        pattern = "".join("?1" if c in VOWELS else "?2" for c in word)
        patterns_by_length.setdefault(length, Counter())[pattern] += 1
    return by_length, patterns_by_length


def bold_row(ws, row_idx):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)


def bold_first_column(ws, start_row=1):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                             min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)


def write_workbook(by_length, patterns_by_length, output_path):
    lengths = sorted(by_length.keys())
    if not lengths:
        raise SystemExit("No words matched the length range — nothing to write.")

    max_vowels = max((max(c.keys()) for c in by_length.values()), default=0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vowel Analysis"

    ws.append(["Word Length"] + [f"{L} characters" for L in lengths])
    ws.append(["Total Words"] + [sum(by_length[L].values()) for L in lengths])

    for v in range(max_vowels + 1):
        row = [v]
        for L in lengths:
            total = sum(by_length[L].values())
            count = by_length[L].get(v, 0)
            row.append(round(count / total * 100, 8) if total else 0)
        ws.append(row)

    bold_row(ws, 1)
    bold_row(ws, 2)
    bold_first_column(ws, start_row=3)

    for L in lengths:
        sheet = wb.create_sheet(title=f"Length_{L}")
        sheet.append(["Pattern", "Count", "Percentage"])
        total = sum(patterns_by_length[L].values())
        for pattern, count in sorted(patterns_by_length[L].items(),
                                     key=lambda x: -x[1]):
            sheet.append([pattern, count, round(count / total * 100, 4)])
        bold_row(sheet, 1)

    wb.save(output_path)


def write_csv(by_length, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("Word Length,Vowel Count,Count,Total Words,Percentage\n")
        for L in sorted(by_length.keys()):
            total = sum(by_length[L].values())
            for v in sorted(by_length[L].keys()):
                count = by_length[L][v]
                pct = count / total * 100 if total else 0
                f.write(f"{L},{v},{count},{total},{pct:.2f}%\n")


def main():
    args = parse_args()
    print(f"Loading {args.input}...")
    words, skipped = load_words(args.input, args.min, args.max)
    print(f"  kept {len(words):,} words "
          f"(skipped {skipped:,} out-of-range / non-alpha)")

    print("Analyzing...")
    by_length, patterns_by_length = analyze(words)

    lengths = sorted(by_length.keys())
    max_vowels = max((max(c.keys()) for c in by_length.values()), default=0)
    total_patterns = sum(len(p) for p in patterns_by_length.values())

    write_workbook(by_length, patterns_by_length, args.output)

    csv_path = os.path.splitext(args.output)[0] + "_vowel_counts.csv"
    write_csv(by_length, csv_path)

    print(f"\nWrote {args.output}")
    print(f"Wrote {csv_path}")
    print(f"  lengths: {lengths[0]}..{lengths[-1]}  "
          f"max vowels: {max_vowels}  "
          f"sheets: Vowel Analysis + {len(lengths)} Length_* "
          f"({total_patterns:,} unique patterns)")


if __name__ == "__main__":
    main()
